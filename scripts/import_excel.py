"""
Import Excel tickets to Supabase
Run: python scripts/import_excel.py
"""
import openpyxl
import uuid
import json
import urllib.request
import urllib.error
from datetime import datetime, timezone

EXCEL_PATH = r'C:\Users\USER\Downloads\Test pack reject log (1).xlsx'
SUPABASE_URL = 'https://vdnbavacjilmgxijarkv.supabase.co'
SERVICE_ROLE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZkbmJhdmFjamlsbWd4aWphcmt2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjA4MTU4NiwiZXhwIjoyMDk3NjU3NTg2fQ.yCiTna5SX4Dmi3DcEx4ypYKStVK5G0-DLLbfBA0UWh0'
IGOR_ID = '48a2756a-3045-4116-9a6a-4d51c4b10614'

# Manual contractor overrides for rows without contractor
CONTRACTOR_OVERRIDES = {
    'TP-10-040-P-001-012': 'TMT',
    'TP-10-029-P-004-001': 'TMT',
    'TP-10-041-P-001-006': 'TMT',
    'TP-10-007-P-003-002': 'TMT',
    'TP-10-009-P-006-002': 'TMT',
}

def normalize_contractor(val):
    if not val:
        return None
    val = str(val).strip()
    if val.upper().startswith('EBS'):
        return 'EBS'
    if val.upper().startswith('TMT'):
        return 'TMT'
    return None

def normalize_status(val):
    if not val:
        return None
    val = str(val).strip()
    if val == 'Reject':
        return 'פתוח'
    if val == 'Accept':
        return 'סגור'
    return None

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Sheet1']
rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

tickets = []
skipped = 0

for row in rows:
    num, date, che, ticket_number, status_raw, issues, contractor_raw = row[0], row[1], row[2], row[3], row[4], row[5], row[6]

    # Skip rows without number or ticket_number
    if not num or not ticket_number:
        skipped += 1
        continue

    status = normalize_status(status_raw)
    if not status:
        skipped += 1
        continue

    contractor = normalize_contractor(contractor_raw)
    if not contractor:
        contractor = CONTRACTOR_OVERRIDES.get(str(ticket_number).strip())
    if not contractor:
        skipped += 1
        continue

    # Date
    if isinstance(date, datetime):
        opened_at = date.replace(tzinfo=timezone.utc).isoformat()
    else:
        opened_at = datetime.now(timezone.utc).isoformat()

    # Description
    description = str(issues).strip() if issues else ''

    # CHE name
    che_name = str(che).strip() if che else 'Igor Ositchansky'
    # Take first name if multiple
    if '\n' in che_name:
        che_name = che_name.split('\n')[0].strip()

    ticket = {
        'id': str(uuid.uuid4()),
        'ticket_number': str(ticket_number).strip(),
        'contractor': contractor,
        'status': status,
        'priority': 'בינונית',
        'description': description,
        'test_phase': None,
        'target_date': None,
        'test_date': None,
        'notes': '',
        'opened_at': opened_at,
        'created_at': opened_at,
        'updated_at': opened_at,
        'closed_at': opened_at if status == 'סגור' else None,
        'created_by_name': che_name,
        'created_by_id': IGOR_ID,
        'assigned_to_id': None,
        'assigned_to_name': None,
        'chat_messages': [],
        'status_history': [],
    }
    tickets.append(ticket)

print(f'Prepared {len(tickets)} tickets, skipped {skipped}')

# Insert in batches of 50
BATCH = 50
headers = {
    'Content-Type': 'application/json',
    'apikey': SERVICE_ROLE_KEY,
    'Authorization': f'Bearer {SERVICE_ROLE_KEY}',
    'Prefer': 'return=minimal',
}

success = 0
for i in range(0, len(tickets), BATCH):
    batch = tickets[i:i+BATCH]
    data = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(
        f'{SUPABASE_URL}/rest/v1/tickets',
        data=data,
        headers=headers,
        method='POST'
    )
    try:
        with urllib.request.urlopen(req) as resp:
            success += len(batch)
            print(f'Inserted batch {i//BATCH + 1}: {len(batch)} tickets (total {success})')
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f'ERROR batch {i//BATCH + 1}: {e.code} {body}')

print(f'\nDone! {success}/{len(tickets)} tickets imported.')
